# UbiqoCR Scraper - AutoMercado
# Angular platform - scroll infinito
# Uso: python scraper_automercado.py

import asyncio
import re
import sys
import json
import random
from datetime import datetime
from collections import Counter

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Instala: python -m pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instala: python -m pip install openpyxl")
    sys.exit(1)

NEGOCIO  = "AutoMercado"
URL_BASE = "https://automercado.cr"

CATEGORIAS = [
    ("https://automercado.cr/categorias/abarrotes",                       "Abarrotes"),
    ("https://automercado.cr/categorias/bebes-y-ninos",                   "Bebes y Ninos"),
    ("https://automercado.cr/categorias/bebidas-alcoholicas",             "Bebidas Alcoholicas"),
    ("https://automercado.cr/categorias/bebidas-no-alcoholicas",          "Bebidas No Alcoholicas"),
    ("https://automercado.cr/categorias/carnes-y-pescado",                "Carnes y Pescado"),
    ("https://automercado.cr/categorias/comidas-preparadas",              "Comidas Preparadas"),
    ("https://automercado.cr/categorias/congelados-y-refrigerados",       "Congelados y Refrigerados"),
    ("https://automercado.cr/categorias/cuidado-personal-y-belleza",      "Cuidado Personal y Belleza"),
    ("https://automercado.cr/categorias/frutas-y-verduras",               "Frutas y Verduras"),
    ("https://automercado.cr/categorias/lacteos-y-embutidos",             "Lacteos y Embutidos"),
    ("https://automercado.cr/categorias/limpieza-y-articulos-desechables","Limpieza y Desechables"),
    ("https://automercado.cr/categorias/mascotas",                        "Mascotas"),
    ("https://automercado.cr/categorias/panaderia-reposteria-y-tortillas","Panaderia y Reposteria"),
    ("https://automercado.cr/categorias/snack-y-golosina",                "Snacks y Golosinas"),
    ("https://automercado.cr/categorias/tienda-y-hogar",                  "Tienda y Hogar"),
]


def limpiar_precio(texto):
    if not texto:
        return None
    limpio = re.sub(r'[^\d.,]', '', texto.strip().split('\n')[0])
    if not limpio:
        return None
    if '.' in limpio and ',' in limpio:
        if limpio.rfind('.') > limpio.rfind(','):
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace('.', '').replace(',', '.')
    elif ',' in limpio:
        partes = limpio.split(',')
        if len(partes) == 2 and len(partes[1]) == 3:
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace(',', '.')
    elif '.' in limpio:
        partes = limpio.split('.')
        if len(partes) >= 2 and len(partes[-1]) == 3:
            limpio = limpio.replace('.', '')
    try:
        r = float(limpio)
        return r if r > 0 else None
    except ValueError:
        return None


def limpiar_texto(s):
    if not isinstance(s, str):
        return s
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s).strip()


async def extraer_productos(page, cat_nombre, vistos):
    nuevos = []
    try:
        items = await page.evaluate("""() => {
            const results = [];
            const cards = document.querySelectorAll('.card-product');
            for (const card of cards) {
                // Nombre: span sin clase
                let nombre = '';
                const spans = card.querySelectorAll('span');
                for (const sp of spans) {
                    const t = sp.textContent.trim();
                    if (!sp.className && t.length > 3 && t.length < 120) {
                        nombre = t;
                        break;
                    }
                }
                if (!nombre) continue;

                // Precio
                const priceEl = card.querySelector('[class*="currency"], [class*="price"], [class*="Price"]');
                const precio = priceEl ? priceEl.textContent.trim() : '';

                // Imagen
                const img = card.querySelector('img');
                const imagen = img ? (img.src || img.dataset.src || '') : '';

                // Link
                const link = card.querySelector('a');
                const href = link ? link.href : '';

                results.push({ nombre, precio, imagen, href });
            }
            return results;
        }""")

        for item in items:
            nombre = limpiar_texto(item.get('nombre', ''))
            if not nombre or len(nombre) < 3:
                continue
            precio = limpiar_precio(item.get('precio', ''))
            imagen = limpiar_texto(item.get('imagen', ''))
            href   = limpiar_texto(item.get('href', ''))

            clave = href if href else nombre.lower()
            if clave in vistos:
                continue
            vistos.add(clave)

            nuevos.append({
                "nombre":    nombre,
                "precio":    precio,
                "moneda":    "CRC",
                "categoria": cat_nombre,
                "negocio":   NEGOCIO,
                "url":       href,
                "imagen":    imagen,
            })
    except Exception as e:
        print(f"      error extrayendo: {str(e)[:40]}")
    return nuevos


async def scrapear():
    productos = []
    vistos    = set()
    SEP = "=" * 60

    print(f"\n{SEP}")
    print(f"  {NEGOCIO} - scroll infinito - {len(CATEGORIAS)} categorias")
    print(f"{SEP}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox",
                  "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
            locale="es-CR",
        )
        await context.route(
            "**/*.{woff,woff2,ttf,eot}",
            lambda route: route.abort()
        )
        page = await context.new_page()

        for url_cat, cat_nombre in CATEGORIAS:
            print(f"\n  [{cat_nombre}]")

            try:
                await page.goto(url_cat, wait_until="domcontentloaded", timeout=60000)
                await asyncio.sleep(4)
            except Exception as e:
                print(f"    error cargando: {str(e)[:50]}")
                continue

            # Esperar que carguen los productos
            try:
                await page.wait_for_selector('.card-product', timeout=15000)
            except PWTimeout:
                print("    sin productos")
                continue

            # Scroll infinito
            scroll_sin_cambio = 0
            ultimo_total = len(productos)
            scroll_count = 0

            while scroll_sin_cambio < 4:
                nuevos = await extraer_productos(page, cat_nombre, vistos)
                productos.extend(nuevos)

                total_actual = len(productos)
                cambio = total_actual - ultimo_total
                print(f"    scroll {scroll_count:>3} | +{cambio:>3} | total: {total_actual:>6}")

                if cambio == 0:
                    scroll_sin_cambio += 1
                else:
                    scroll_sin_cambio = 0
                    ultimo_total = total_actual

                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await asyncio.sleep(random.uniform(2.0, 3.0))
                scroll_count += 1

                if scroll_count > 200:
                    break

            print(f"    -> Fin ({scroll_count} scrolls)")

        await browser.close()

    sin_p = sum(1 for p in productos if p["precio"] is None)
    print(f"\n  TOTAL {NEGOCIO}: {len(productos)} productos ({sin_p} sin precio)")
    return productos


def exportar_excel(productos):
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_{NEGOCIO}_{fecha}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    AZUL  = "003087"
    CLARO = "E8EEF7"
    GRIS  = "F8FAFC"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hx = PatternFill("solid", fgColor=AZUL)
    ha = Alignment(horizontal="center", vertical="center")
    ax = PatternFill("solid", fgColor=CLARO)
    gx = PatternFill("solid", fgColor=GRIS)
    na = Alignment(vertical="center")
    ra = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="E2E8F0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols   = ["Nombre del producto", "Precio", "Moneda", "Categoria", "Negocio", "URL", "Imagen URL"]
    widths = [55, 14, 9, 25, 14, 60, 60]

    for ci, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = hf; cell.fill = hx
        cell.alignment = ha; cell.border = brd
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for i, prod in enumerate(productos, 1):
        fila = i + 1
        fill = ax if i % 2 == 0 else gx
        vals = [prod.get("nombre",""), prod.get("precio"),
                prod.get("moneda","CRC"), prod.get("categoria",""),
                prod.get("negocio",""), prod.get("url",""), prod.get("imagen","")]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, str):
                v = limpiar_texto(v)
            cell = ws.cell(row=fila, column=ci, value=v)
            cell.border = brd; cell.fill = fill
            if ci == 2 and isinstance(v, float):
                cell.number_format = "#,##0.00"; cell.alignment = ra
            else:
                cell.alignment = na
        ws.row_dimensions[fila].height = 16

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = f"{NEGOCIO} - {len(productos)} productos"
    ws2["A1"].font = Font(bold=True, size=13, color=AZUL)
    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    for ci, t in enumerate(["Categoria", "Productos", "Con precio"], 1):
        c = ws2.cell(row=4, column=ci, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)

    cnt = Counter(p["categoria"] for p in productos)
    cpr = Counter(p["categoria"] for p in productos if p["precio"] is not None)
    for ri, (cat, n) in enumerate(sorted(cnt.items()), 5):
        ws2.cell(row=ri, column=1, value=cat)
        ws2.cell(row=ri, column=2, value=n)
        ws2.cell(row=ri, column=3, value=cpr.get(cat, 0))

    tf = 5 + len(cnt)
    ws2.cell(row=tf, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=tf, column=2, value=len(productos)).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(nombre_f)
    print(f"  Excel: {nombre_f}")
    return nombre_f


def exportar_json(productos):
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_{NEGOCIO}_{fecha}.json"
    with open(nombre_f, "w", encoding="utf-8") as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)
    print(f"  JSON:  {nombre_f}")
    return nombre_f


async def main():
    print(f"\nUbiqoCR Scraper - {NEGOCIO}")
    print(f"Inicio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    prods = await scrapear()
    if prods:
        exportar_excel(prods)
        exportar_json(prods)
    else:
        print("Sin productos")

    print(f"\nFin: {datetime.now().strftime('%H:%M:%S')}\n")


if __name__ == "__main__":
    asyncio.run(main())
