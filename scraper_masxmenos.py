# UbiqoCR Scraper - MasxMenos
# VTEX API con IDs de categoria
# Uso: python scraper_masxmenos.py

import asyncio
import re
import ssl
import json
import random
import sys
from datetime import datetime
from collections import Counter

try:
    import aiohttp
except ImportError:
    print("Instala: python312 -m pip install aiohttp")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instala: python312 -m pip install openpyxl")
    sys.exit(1)

NEGOCIO  = "MasxMenos"
URL_BASE = "https://www.masxmenos.cr"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json",
}

# (categoria_id, nombre_categoria)
# Usando IDs de categorias PRINCIPALES que sí devuelven productos
CATEGORIAS = [
    (1,  "Jugos y Bebidas"),
    (2,  "Cervezas Vinos y Licores"),
    (3,  "Carnes y Pescados"),
    (6,  "Higiene y Belleza"),
    (8,  "Bebes y Ninos"),
    (9,  "Alimentos Congelados"),
    (10, "Lacteos"),
    (11, "Abarrotes"),
    (13, "Limpieza"),
    (14, "Farmacia"),
    (15, "Frutas y Verduras"),
    (16, "Panaderia y Tortilleria"),
    (21653, "Mascotas"),
]


def limpiar_precio(precio_raw):
    if isinstance(precio_raw, (int, float)):
        return float(precio_raw)
    if isinstance(precio_raw, str):
        limpio = re.sub(r'[^\d.]', '', precio_raw)
        try:
            return float(limpio)
        except ValueError:
            pass
    return None


def limpiar_texto(s):
    if not isinstance(s, str):
        return s
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s).strip()


async def get_productos_cat(session, cat_id, subcat_nombre, cat_nombre, vistos):
    nuevos = []
    desde  = 0
    paso   = 50

    while True:
        url = (
            f"{URL_BASE}/api/catalog_system/pub/products/search"
            f"?fq=C:/{cat_id}/"
            f"&_from={desde}&_to={desde + paso - 1}"
            f"&O=OrderByReleaseDateDESC"
        )
        try:
            async with session.get(url, headers=HEADERS, timeout=aiohttp.ClientTimeout(total=20)) as r:
                if r.status not in (200, 206):
                    break
                items = await r.json(content_type=None)
                if not items:
                    break

                resources = r.headers.get('resources', '')
                try:
                    total = int(resources.split('/')[-1])
                except Exception:
                    total = 0

                for item in items:
                    try:
                        nombre = limpiar_texto(item.get('productName', '') or item.get('name', ''))
                        if not nombre or len(nombre) < 3:
                            continue

                        # Precio
                        precio = None
                        try:
                            skus = item.get('items', [])
                            if skus:
                                sellers = skus[0].get('sellers', [])
                                if sellers:
                                    offer = sellers[0].get('commertialOffer', {})
                                    p = offer.get('Price') or offer.get('ListPrice')
                                    if p:
                                        precio = float(p)
                        except Exception:
                            pass

                        # Imagen
                        imagen = ""
                        try:
                            skus = item.get('items', [])
                            if skus:
                                imgs = skus[0].get('images', [])
                                if imgs:
                                    imagen = imgs[0].get('imageUrl', '')
                        except Exception:
                            pass

                        href = f"{URL_BASE}/{item.get('linkText', '')}/p"
                        prod_id = str(item.get('productId', ''))

                        clave = prod_id if prod_id else nombre.lower()
                        if clave in vistos:
                            continue
                        vistos.add(clave)

                        nuevos.append({
                            "nombre":    nombre,
                            "precio":    precio,
                            "moneda":    "CRC",
                            "categoria": cat_nombre,
                            "negocio":   NEGOCIO,
                            "url":       limpiar_texto(href),
                            "imagen":    limpiar_texto(imagen),
                        })
                    except Exception:
                        continue

                desde += paso
                if desde >= total or len(items) < paso:
                    break

                await asyncio.sleep(random.uniform(0.2, 0.4))

        except Exception as e:
            print(f"      error: {str(e)[:40]}")
            break

    return nuevos


async def scrapear():
    productos = []
    vistos    = set()
    SEP = "=" * 60

    print(f"\n{SEP}")
    print(f"  {NEGOCIO} - VTEX API - {len(CATEGORIAS)} subcategorias")
    print(f"{SEP}")

    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    connector = aiohttp.TCPConnector(ssl=ssl_ctx, limit=5)

    async with aiohttp.ClientSession(connector=connector) as session:
        cat_actual = ""
        for cat_id, cat_nombre in CATEGORIAS:
            print(f"\n[{cat_nombre}]", flush=True)
            nuevos = await get_productos_cat(session, cat_id, cat_nombre, cat_nombre, vistos)
            productos.extend(nuevos)
            print(f" {len(nuevos)} | total: {len(productos)}")

            await asyncio.sleep(random.uniform(0.3, 0.6))

    sin_p = sum(1 for p in productos if p["precio"] is None)
    print(f"\n  TOTAL {NEGOCIO}: {len(productos)} productos ({sin_p} sin precio)")
    return productos


def exportar_excel(productos):
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_{NEGOCIO}_{fecha}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    COLOR = "FF6B00"
    CLARO = "FFF3E0"
    GRIS  = "F8FAFC"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hx = PatternFill("solid", fgColor=COLOR)
    ha = Alignment(horizontal="center", vertical="center")
    ax = PatternFill("solid", fgColor=CLARO)
    gx = PatternFill("solid", fgColor=GRIS)
    na = Alignment(vertical="center")
    ra = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="E2E8F0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols   = ["Nombre del producto", "Precio", "Moneda", "Categoria", "Negocio", "URL", "Imagen URL"]
    widths = [55, 14, 9, 25, 12, 60, 60]

    for ci, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = hf; cell.fill = hx
        cell.alignment = ha; cell.border = brd
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for i, prod in enumerate(productos, 1):
        fila = i + 1
        fill = ax if i % 2 == 0 else gx
        vals = [prod.get("nombre",""), prod.get("precio"),
                prod.get("moneda","CRC"), prod.get("categoria",""),
                prod.get("negocio",""), prod.get("url",""), prod.get("imagen","")]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, str):
                v = limpiar_texto(v)
            cell = ws.cell(row=fila, column=ci, value=v)
            cell.border = brd; cell.fill = fill
            if ci == 2 and isinstance(v, float):
                cell.number_format = "#,##0.00"; cell.alignment = ra
            else:
                cell.alignment = na
        ws.row_dimensions[fila].height = 16

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = f"{NEGOCIO} - {len(productos)} productos"
    ws2["A1"].font = Font(bold=True, size=13, color=COLOR)
    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    for ci, t in enumerate(["Categoria", "Productos", "Con precio"], 1):
        c = ws2.cell(row=4, column=ci, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLOR)

    cnt = Counter(p["categoria"] for p in productos)
    cpr = Counter(p["categoria"] for p in productos if p["precio"] is not None)
    for ri, (cat, n) in enumerate(sorted(cnt.items()), 5):
        ws2.cell(row=ri, column=1, value=cat)
        ws2.cell(row=ri, column=2, value=n)
        ws2.cell(row=ri, column=3, value=cpr.get(cat, 0))

    tf = 5 + len(cnt)
    ws2.cell(row=tf, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=tf, column=2, value=len(productos)).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(nombre_f)
    print(f"  Excel: {nombre_f}")
    return nombre_f


def exportar_json(productos):
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_{NEGOCIO}_{fecha}.json"
    with open(nombre_f, "w", encoding="utf-8") as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)
    print(f"  JSON:  {nombre_f}")
    return nombre_f


async def main():
    print(f"\nUbiqoCR Scraper - {NEGOCIO}")
    print(f"Inicio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    prods = await scrapear()
    if prods:
        exportar_excel(prods)
        exportar_json(prods)
    else:
        print("Sin productos")

    print(f"\nFin: {datetime.now().strftime('%H:%M:%S')}\n")


if __name__ == "__main__":
    asyncio.run(main())
