# UbiqoCR Scraper - El Lagar
# Paginacion por clic en boton siguiente (JavaScript dinamico)
# Uso: python scraper_lagar.py

import asyncio
import re
import sys
import random
from datetime import datetime
from collections import Counter

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Instala: python -m pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instala: python -m pip install openpyxl")
    sys.exit(1)

CATEGORIAS_LAGAR = [
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/7/0/herramientas",       "Herramientas"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/10/0/pinturas",          "Pinturas"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/21/0/iluminaci%C3%B3n",  "Iluminacion"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/8/0/art%C3%ADculos-para-el-hogar", "Hogar"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/15/0/jardiner%C3%ADa",   "Jardineria"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/1/0/pisos-ba%C3%B1os-y-cocinas", "Pisos Banos Cocinas"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/17/0/automotriz",        "Automotriz"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/11/0/fontaner%C3%ADa",   "Plomeria"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/4/0/el%C3%A9ctrico",     "Electrico"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/2/0/aceros",             "Aceros"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/9/0/construcci%C3%B3n",  "Construccion"),
    ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/18/0/exteriores-y-camping", "Exteriores"),
]


def limpiar_precio(texto):
    if not texto:
        return None
    limpio = re.sub(r'[^\d.,]', '', texto.strip())
    if not limpio:
        return None
    tiene_punto = '.' in limpio
    tiene_coma  = ',' in limpio
    if tiene_punto and tiene_coma:
        if limpio.rfind('.') > limpio.rfind(','):
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace('.', '').replace(',', '.')
    elif tiene_punto:
        partes = limpio.split('.')
        if len(partes) >= 2 and len(partes[-1]) == 3:
            limpio = limpio.replace('.', '')
        elif len(partes) > 2:
            limpio = limpio.replace('.', '')
    elif tiene_coma:
        partes = limpio.split(',')
        if len(partes) == 2 and len(partes[1]) == 3:
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace(',', '.')
    try:
        r = float(limpio)
        return r if r > 0 else None
    except ValueError:
        return None


def limpiar_texto(s):
    if not isinstance(s, str):
        return s
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s).strip()


def extraer_precio_de_texto(texto):
    match = re.search(r'[¢$]?\s*([\d][.,\d]+)', texto)
    if match:
        return limpiar_precio(match.group(1))
    return None


async def extraer_productos_pagina(page, cat_nombre, vistos):
    nuevos = []
    links = await page.query_selector_all('a[href*="DetalleArticulo"]')
    for link in links:
        try:
            # Obtener el texto completo del contenedor del producto
            parent_txt = await link.evaluate(
                "el => el.parentElement ? el.parentElement.innerText : ''"
            )

            # Separar lineas y limpiar
            lines = [l.strip() for l in parent_txt.split('\n') if l.strip()]

            # La primera linea que no tenga precio ni "Agregar" es el nombre
            nombre = ""
            precio_txt = ""
            for line in lines:
                if 'Agregar' in line or 'agregar' in line:
                    continue
                # Si la linea tiene simbolo de colones o patron de precio, es el precio
                if re.search(r'[¢$]|IVA', line):
                    if not precio_txt:
                        precio_txt = line
                else:
                    if not nombre and len(line) > 3:
                        nombre = limpiar_texto(line)

            # Si no encontro nombre limpio, usar innerText del link directamente
            if not nombre:
                nombre = limpiar_texto((await link.inner_text()).strip())
                # Quitar precio pegado al nombre si lo tiene
                nombre = re.sub(r'[¢$][\d.,]+.*$', '', nombre).strip()
                nombre = re.sub(r'\d{1,3}[.,]\d{3}.*$', '', nombre).strip()

            if not nombre or len(nombre) < 3:
                continue

            # Extraer precio numerico
            precio = extraer_precio_de_texto(precio_txt) if precio_txt else None
            if precio is None:
                precio = extraer_precio_de_texto(parent_txt)

            # URL
            href = limpiar_texto(await link.get_attribute("href") or "")
            if href and not href.startswith("http"):
                href = "https://www.ellagar.com" + href

            # Deduplicar por URL (mas confiable que nombre)
            clave = href if href else nombre.lower()
            if clave in vistos:
                continue
            vistos.add(clave)

            nuevos.append({
                "nombre":    nombre,
                "precio":    precio,
                "moneda":    "CRC",
                "categoria": cat_nombre,
                "url":       href,
            })
        except Exception:
            continue
    return nuevos


async def scrapear():
    productos = []
    vistos    = set()

    print(f"\n{'='*60}")
    print(f"  El Lagar - {len(CATEGORIAS_LAGAR)} categorias")
    print(f"{'='*60}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox",
                  "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 900},
            locale="es-CR",
        )
        page = await context.new_page()

        await page.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,woff,woff2,ttf,eot,ico}",
            lambda route: route.abort()
        )

        for url_cat, cat_nombre in CATEGORIAS_LAGAR:
            print(f"\n  [{cat_nombre}]")

            try:
                resp = await page.goto(url_cat, wait_until="domcontentloaded", timeout=30000)
                if resp and resp.status in (404, 403, 410):
                    print(f"    HTTP {resp.status}, saltando")
                    continue

                await page.wait_for_selector('a[href*="DetalleArticulo"]', timeout=15000)

            except Exception as e:
                print(f"    Error cargando categoria: {str(e)[:50]}")
                continue

            pagina = 1
            while True:
                print(f"    p{pagina:>3} -> ", end="", flush=True)

                # Esperar que los productos esten listos
                await asyncio.sleep(0.5)

                nuevos = await extraer_productos_pagina(page, cat_nombre, vistos)
                productos.extend(nuevos)
                print(f"{len(nuevos):>3} nuevos | total: {len(productos):>6}")

                # Buscar boton siguiente — El Lagar usa un span/a con el caracter >
                # Intentar varios selectores posibles
                btn_sig = None

                # Selector exacto del boton siguiente de El Lagar
                # class="rc-pagination-next", aria-label="next page"
                try:
                    btn_sig = await page.query_selector('li.rc-pagination-next:not([aria-disabled="true"]) button')
                    if not btn_sig:
                        btn_sig = await page.query_selector('button[aria-label="next page"]')
                    if not btn_sig:
                        # Verificar si esta deshabilitado
                        disabled = await page.query_selector('li.rc-pagination-next[aria-disabled="true"]')
                        if disabled:
                            btn_sig = None  # ultima pagina
                except Exception:
                    btn_sig = None

                if not btn_sig:
                    print(f"    -> Fin de categoria ({pagina} paginas)")
                    break

                # Guardar el primer producto de la pagina actual para detectar cambio
                primer_antes = await page.evaluate(
                    "() => { const links = document.querySelectorAll('a[href*=\"DetalleArticulo\"]'); return links.length > 0 ? links[0].href : ''; }"
                )

                # Hacer clic en siguiente
                try:
                    await btn_sig.click()
                    # Esperar que cambie el contenido
                    await asyncio.sleep(1.5)
                    await page.wait_for_selector('a[href*="DetalleArticulo"]', timeout=10000)

                    # Verificar que realmente cambio la pagina
                    primer_despues = await page.evaluate(
                        "() => { const links = document.querySelectorAll('a[href*=\"DetalleArticulo\"]'); return links.length > 0 ? links[0].href : ''; }"
                    )

                    if primer_despues == primer_antes:
                        print(f"    -> Pagina no cambio, fin de categoria")
                        break

                except Exception as e:
                    print(f"    -> Error en clic siguiente: {str(e)[:40]}")
                    break

                pagina += 1
                await asyncio.sleep(random.uniform(0.5, 1.2))

        await browser.close()

    sin_p = sum(1 for p in productos if p["precio"] is None)
    print(f"\n  TOTAL El Lagar: {len(productos)} productos ({sin_p} sin precio)")
    return productos


def exportar_excel(productos):
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_ElLagar_{fecha}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    AZUL  = "C00404"
    CLARO = "FFF0F0"
    GRIS  = "F8FAFC"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hx = PatternFill("solid", fgColor=AZUL)
    ha = Alignment(horizontal="center", vertical="center")
    ax = PatternFill("solid", fgColor=CLARO)
    gx = PatternFill("solid", fgColor=GRIS)
    na = Alignment(vertical="center")
    ra = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="E2E8F0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols   = ["Nombre del producto", "Precio", "Moneda", "Categoria", "URL"]
    widths = [55, 14, 9, 25, 60]

    for ci, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = hf; cell.fill = hx
        cell.alignment = ha; cell.border = brd
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for i, prod in enumerate(productos, 1):
        fila = i + 1
        fill = ax if i % 2 == 0 else gx
        vals = [prod.get("nombre",""), prod.get("precio"),
                prod.get("moneda","CRC"), prod.get("categoria",""), prod.get("url","")]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, str):
                v = limpiar_texto(v)
            cell = ws.cell(row=fila, column=ci, value=v)
            cell.border = brd; cell.fill = fill
            if ci == 2 and isinstance(v, float):
                cell.number_format = "#,##0.00"; cell.alignment = ra
            else:
                cell.alignment = na
        ws.row_dimensions[fila].height = 16

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = f"El Lagar - {len(productos)} productos"
    ws2["A1"].font = Font(bold=True, size=13, color=AZUL)
    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    for ci, t in enumerate(["Categoria", "Productos", "Con precio"], 1):
        c = ws2.cell(row=4, column=ci, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)

    cnt = Counter(p["categoria"] for p in productos)
    cpr = Counter(p["categoria"] for p in productos if p["precio"] is not None)
    for ri, (cat, n) in enumerate(sorted(cnt.items()), 5):
        ws2.cell(row=ri, column=1, value=cat)
        ws2.cell(row=ri, column=2, value=n)
        ws2.cell(row=ri, column=3, value=cpr.get(cat, 0))

    tf = 5 + len(cnt)
    ws2.cell(row=tf, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=tf, column=2, value=len(productos)).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(nombre_f)
    print(f"  Excel: {nombre_f}")
    return nombre_f


async def main():
    print(f"\nUbiqoCR Scraper - El Lagar")
    print(f"Inicio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    prods = await scrapear()
    if prods:
        exportar_excel(prods)
    else:
        print("Sin productos extraidos")

    print(f"\nFin: {datetime.now().strftime('%H:%M:%S')}\n")


if __name__ == "__main__":
    asyncio.run(main())
