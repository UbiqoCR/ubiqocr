# UbiqoCR Web Scraper v4
# URLs exactas por sitio
#
# Uso:
#   python scraper.py --sitio epa
#   python scraper.py --sitio lagar
#   python scraper.py --sitio brenes
#   python scraper.py --sitio novex
#   python scraper.py --sitio monge
#   python scraper.py --sitio gollo

import asyncio
import re
import sys
import random
from datetime import datetime
from collections import Counter

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Instala: python -m pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instala: python -m pip install openpyxl")
    sys.exit(1)


# ==============================================================
# CONFIGURACION
# Cada categoria es (url_completa, nombre_categoria)
# ==============================================================
SITIOS = {

    "epa": {
        "nombre": "EPA",
        "tipo": "magento",
        "categorias": [
            ("https://cr.epaenlinea.com/productos.html?cat=13829", "Ferreteria"),
            ("https://cr.epaenlinea.com/productos.html?cat=13832", "Herramientas"),
            ("https://cr.epaenlinea.com/productos.html?cat=13835", "Pinturas"),
            ("https://cr.epaenlinea.com/productos.html?cat=13838", "Electrico"),
            ("https://cr.epaenlinea.com/productos.html?cat=13841", "Plomeria"),
            ("https://cr.epaenlinea.com/productos.html?cat=13844", "Construccion"),
            ("https://cr.epaenlinea.com/productos.html?cat=13847", "Hogar"),
            ("https://cr.epaenlinea.com/productos.html?cat=13850", "Bano"),
            ("https://cr.epaenlinea.com/productos.html?cat=13853", "Cocina"),
            ("https://cr.epaenlinea.com/productos.html?cat=13856", "Jardin"),
            ("https://cr.epaenlinea.com/productos.html?cat=13859", "Piso y Pared"),
            ("https://cr.epaenlinea.com/productos.html?cat=13862", "Iluminacion"),
            ("https://cr.epaenlinea.com/productos.html?cat=13865", "Cerrajeria"),
            ("https://cr.epaenlinea.com/productos.html?cat=13868", "Adhesivos"),
            ("https://cr.epaenlinea.com/productos.html?cat=13871", "Limpieza"),
            ("https://cr.epaenlinea.com/productos.html?cat=13874", "Seguridad"),
            ("https://cr.epaenlinea.com/productos.html?cat=13877", "Organizacion"),
            ("https://cr.epaenlinea.com/productos.html?cat=13880", "Automotriz"),
        ],
        "selector_prod":  ".product-item-info",
        "selector_nombre": ".product-item-name a",
        "selector_precio": "span.price",
        "selector_sig":    "a.action.next",
        "paginacion": "?p=",   # se agrega al final de la URL base
    },

    "lagar": {
        "nombre": "El Lagar",
        "tipo": "lagar",
        "categorias": [
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/7/0/herramientas",      "Herramientas"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/10/0/pinturas",         "Pinturas"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/21/0/iluminaci%C3%B3n","Iluminacion"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/8/0/art%C3%ADculos-para-el-hogar", "Hogar"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/15/0/jardiner%C3%ADa", "Jardineria"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/1/0/pisos-ba%C3%B1os-y-cocinas", "Pisos Banos Cocinas"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/17/0/automotriz",       "Automotriz"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/11/0/fontaner%C3%ADa", "Plomeria"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/4/0/el%C3%A9ctrico",   "Electrico"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/2/0/aceros",            "Aceros"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/9/0/construcci%C3%B3n","Construccion"),
            ("https://www.ellagar.com/ECOMMERCE/CategoriaArticulo/18/0/exteriores-y-camping", "Exteriores"),
        ],
        # El Lagar usa un sistema propio, intentamos varios selectores
        "selector_prod":   ".product-item, .articulo, article, [class*=product], [class*=articulo]",
        "selector_nombre": "a.product-item-name, a[class*=nombre], h3 a, h2 a, .nombre a, a[title]",
        "selector_precio": "span.price, .precio, [class*=price], [class*=precio]",
        "selector_sig":    "a.action.next, a[class*=next], .siguiente a, li.next a",
        "paginacion": "?p=",
    },

    "brenes": {
        "nombre": "Ferreteria Brenes",
        "tipo": "woocommerce",
        "categorias": [
            ("https://ferreteriabrenes.com/categoria/salud-ocupacional/", "Salud Ocupacional"),
            ("https://ferreteriabrenes.com/categoria/pinturas/",          "Pinturas"),
            ("https://ferreteriabrenes.com/categoria/jardineria/jardin/", "Jardineria"),
            ("https://ferreteriabrenes.com/categoria/iluminacion/",       "Iluminacion"),
            ("https://ferreteriabrenes.com/categoria/hogar/",             "Hogar"),
            ("https://ferreteriabrenes.com/categoria/herramientas/",      "Herramientas"),
            ("https://ferreteriabrenes.com/categoria/griferia/",          "Griferia"),
            ("https://ferreteriabrenes.com/categoria/ferreteria/",        "Ferreteria"),
            ("https://ferreteriabrenes.com/categoria/electrico/",         "Electrico"),
            ("https://ferreteriabrenes.com/categoria/construccion/",      "Construccion"),
            ("https://ferreteriabrenes.com/categoria/cerrajeria/",        "Cerrajeria"),
        ],
        # WooCommerce tiene selectores estándar
        "selector_prod":   "li.product, .product-item, .type-product",
        "selector_nombre": "h2.woocommerce-loop-product__title, a.woocommerce-LoopProduct-link h2, h2.product-title, .product-title",
        "selector_precio": "span.price bdi, span.price, .woocommerce-Price-amount bdi, .woocommerce-Price-amount",
        "selector_sig":    "a.next.page-numbers, li.next a, a[aria-label='Next']",
        "paginacion": "/page/",  # WooCommerce: /categoria/x/page/2/
    },

    "novex": {
        "nombre": "Novex",
        "tipo": "magento",
        "categorias": [
            ("https://novex.cr/catalogo/40/Ferreter%C3%ADa.html",              "Ferreteria"),
            ("https://novex.cr/catalogo/41/Herramientas-El%C3%A9ctricas.html", "Herramientas Electricas"),
            ("https://novex.cr/catalogo/42/Herramientas-Manuales.html",        "Herramientas Manuales"),
            ("https://novex.cr/catalogo/43/Fontaner%C3%ADa.html",              "Plomeria"),
            ("https://novex.cr/catalogo/44/Ba%C3%B1os.html",                   "Banos"),
            ("https://novex.cr/catalogo/45/Pintura.html",                      "Pinturas"),
            ("https://novex.cr/catalogo/46/Cerrajer%C3%ADa.html",              "Cerrajeria"),
            ("https://novex.cr/catalogo/47/Torniller%C3%ADa.html",             "Tornilleria"),
            ("https://novex.cr/catalogo/48/Hogar.html",                        "Hogar"),
            ("https://novex.cr/catalogo/49/Iluminaci%C3%B3n.html",             "Iluminacion"),
            ("https://novex.cr/catalogo/50/El%C3%A9ctrico.html",               "Electrico"),
            ("https://novex.cr/catalogo/51/Jardiner%C3%ADa.html",              "Jardineria"),
            ("https://novex.cr/catalogo/52/Automotriz.html",                   "Automotriz"),
            ("https://novex.cr/catalogo/53/Limpieza.html",                     "Limpieza"),
            ("https://novex.cr/catalogo/54/Construcci%C3%B3n.html",            "Construccion"),
        ],
        "selector_prod":   ".product-item-info, .product-item",
        "selector_nombre": ".product-item-name a, .product-name a",
        "selector_precio": "span.price, .price",
        "selector_sig":    "a.action.next",
        "paginacion": "?p=",
    },

    "monge": {
        "nombre": "Tienda Monge",
        "tipo": "magento",
        "categorias": [
            ("https://www.tiendamonge.com/electrodomesticos.html",  "Electrodomesticos"),
            ("https://www.tiendamonge.com/television.html",         "Television"),
            ("https://www.tiendamonge.com/audio.html",              "Audio"),
            ("https://www.tiendamonge.com/celulares.html",          "Celulares"),
            ("https://www.tiendamonge.com/computo.html",            "Computo"),
            ("https://www.tiendamonge.com/refrigeracion.html",      "Refrigeracion"),
            ("https://www.tiendamonge.com/lavado.html",             "Lavado"),
            ("https://www.tiendamonge.com/cocina.html",             "Cocina"),
            ("https://www.tiendamonge.com/muebles.html",            "Muebles"),
            ("https://www.tiendamonge.com/ferreteria.html",         "Ferreteria"),
        ],
        "selector_prod":   ".product-item-info",
        "selector_nombre": ".product-item-name a",
        "selector_precio": "span.price",
        "selector_sig":    "a.action.next",
        "paginacion": "?p=",
    },

    "gollo": {
        "nombre": "Gollo",
        "tipo": "magento",
        "categorias": [
            ("https://www.gollo.com/electrodomesticos.html",  "Electrodomesticos"),
            ("https://www.gollo.com/television.html",         "Television"),
            ("https://www.gollo.com/audio.html",              "Audio"),
            ("https://www.gollo.com/celulares.html",          "Celulares"),
            ("https://www.gollo.com/computo.html",            "Computo"),
            ("https://www.gollo.com/refrigeracion.html",      "Refrigeracion"),
            ("https://www.gollo.com/lavado.html",             "Lavado"),
            ("https://www.gollo.com/cocina.html",             "Cocina"),
            ("https://www.gollo.com/muebles.html",            "Muebles"),
        ],
        "selector_prod":   ".product-item-info",
        "selector_nombre": ".product-item-name a",
        "selector_precio": "span.price",
        "selector_sig":    "a.action.next",
        "paginacion": "?p=",
    },
}


# ==============================================================
# UTILIDADES
# ==============================================================

def limpiar_precio(texto):
    if not texto:
        return None
    limpio = re.sub(r'[^\d.,]', '', texto.strip())
    if not limpio:
        return None
    tiene_punto = '.' in limpio
    tiene_coma  = ',' in limpio
    if tiene_punto and tiene_coma:
        if limpio.rfind('.') > limpio.rfind(','):
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace('.', '').replace(',', '.')
    elif tiene_punto:
        partes = limpio.split('.')
        if len(partes) >= 2 and len(partes[-1]) == 3:
            limpio = limpio.replace('.', '')
        elif len(partes) > 2:
            limpio = limpio.replace('.', '')
    elif tiene_coma:
        partes = limpio.split(',')
        if len(partes) == 2 and len(partes[1]) == 3:
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace(',', '.')
    try:
        r = float(limpio)
        return r if r > 0 else None
    except ValueError:
        return None


def limpiar_texto(s):
    if not isinstance(s, str):
        return s
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s).strip()


def url_pagina(url_base, pagina, estilo_pag):
    if pagina == 1:
        return url_base
    if estilo_pag == "/page/":
        # WooCommerce: insertar /page/N/ antes del ultimo /
        url = url_base.rstrip('/')
        return url + f"/page/{pagina}/"
    else:
        # Magento: ?p=N o &p=N
        sep = "&" if "?" in url_base else "?"
        return url_base + sep + f"p={pagina}"


# ==============================================================
# SCRAPER
# ==============================================================

async def scrapear_sitio(cfg):
    nombre    = cfg["nombre"]
    productos = []
    vistos    = set()
    pag_estilo = cfg.get("paginacion", "?p=")

    print(f"\n{'='*60}")
    print(f"  {nombre}")
    print(f"{'='*60}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-blink-features=AutomationControlled",
            ]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
            locale="es-CR",
            extra_http_headers={"Accept-Language": "es-CR,es;q=0.9"},
        )
        page = await context.new_page()

        await page.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,woff,woff2,ttf,eot,ico}",
            lambda route: route.abort()
        )

        for url_cat, cat_nombre in cfg["categorias"]:
            pagina = 1
            errores = 0
            print(f"\n  [{cat_nombre}]")

            while pagina <= 100:
                url = url_pagina(url_cat, pagina, pag_estilo)
                print(f"    p{pagina:>3} -> ", end="", flush=True)

                try:
                    resp = await page.goto(url, wait_until="domcontentloaded", timeout=25000)

                    if resp and resp.status in (404, 403, 410):
                        print(f"HTTP {resp.status}")
                        break

                    # Esperar selector de productos
                    sel_prod = None
                    for s in cfg["selector_prod"].split(", "):
                        try:
                            await page.wait_for_selector(s.strip(), timeout=8000)
                            sel_prod = s.strip()
                            break
                        except PWTimeout:
                            continue

                    if not sel_prod:
                        print("sin productos")
                        errores += 1
                        if errores >= 3:
                            break
                        continue

                except PWTimeout:
                    print("timeout")
                    errores += 1
                    if errores >= 3:
                        break
                    await asyncio.sleep(2)
                    continue
                except Exception as e:
                    print(f"error: {str(e)[:50]}")
                    errores += 1
                    if errores >= 3:
                        break
                    continue

                errores = 0
                items = await page.query_selector_all(sel_prod)

                if not items:
                    print("vacio")
                    break

                nuevos = 0
                for item in items:
                    try:
                        # Nombre
                        nombre_prod = ""
                        for s in cfg["selector_nombre"].split(", "):
                            el = await item.query_selector(s.strip())
                            if el:
                                nombre_prod = limpiar_texto((await el.inner_text()).strip())
                                if nombre_prod:
                                    break

                        if not nombre_prod:
                            continue

                        # Precio
                        precio_txt = ""
                        for s in cfg["selector_precio"].split(", "):
                            el = await item.query_selector(s.strip())
                            if el:
                                precio_txt = (await el.inner_text()).strip()
                                if precio_txt:
                                    break
                        precio_num = limpiar_precio(precio_txt)

                        # URL
                        href = ""
                        for s in cfg["selector_nombre"].split(", "):
                            el = await item.query_selector(s.strip())
                            if el:
                                href = limpiar_texto(await el.get_attribute("href") or "")
                                if href:
                                    break

                        clave = nombre_prod.lower()
                        if clave in vistos:
                            continue
                        vistos.add(clave)

                        productos.append({
                            "nombre":    nombre_prod,
                            "precio":    precio_num,
                            "moneda":    "CRC",
                            "categoria": cat_nombre,
                            "url":       href,
                        })
                        nuevos += 1

                    except Exception:
                        continue

                print(f"{nuevos:>3} nuevos | total: {len(productos):>5}")

                # Siguiente pagina
                sig = None
                for s in cfg["selector_sig"].split(", "):
                    try:
                        sig = await page.query_selector(s.strip())
                        if sig:
                            break
                    except Exception:
                        continue

                if not sig:
                    break

                pagina += 1
                await asyncio.sleep(random.uniform(1.0, 2.2))

        await browser.close()

    sin_precio = sum(1 for p in productos if p["precio"] is None)
    print(f"\n  OK {nombre}: {len(productos)} productos ({sin_precio} sin precio)")
    return productos


# ==============================================================
# EXPORTAR EXCEL
# ==============================================================

def exportar_excel(productos, sitio_nombre):
    if not productos:
        print(f"  Sin productos: {sitio_nombre}")
        return ""

    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_{sitio_nombre.replace(' ', '_')}_{fecha}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    AZUL  = "2563EB"
    CLARO = "EFF6FF"
    GRIS  = "F8FAFC"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hx = PatternFill("solid", fgColor=AZUL)
    ha = Alignment(horizontal="center", vertical="center")
    ax = PatternFill("solid", fgColor=CLARO)
    gx = PatternFill("solid", fgColor=GRIS)
    na = Alignment(vertical="center")
    ra = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="E2E8F0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols   = ["Nombre del producto", "Precio", "Moneda", "Categoria", "URL"]
    widths = [55, 14, 9, 25, 60]

    for ci, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = hf
        cell.fill = hx
        cell.alignment = ha
        cell.border = brd
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for i, prod in enumerate(productos, 1):
        fila = i + 1
        fill = ax if i % 2 == 0 else gx
        vals = [
            prod.get("nombre", ""),
            prod.get("precio"),
            prod.get("moneda", "CRC"),
            prod.get("categoria", ""),
            prod.get("url", ""),
        ]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, str):
                v = limpiar_texto(v)
            cell = ws.cell(row=fila, column=ci, value=v)
            cell.border = brd
            cell.fill   = fill
            if ci == 2 and isinstance(v, float):
                cell.number_format = "#,##0.00"
                cell.alignment = ra
            else:
                cell.alignment = na
        ws.row_dimensions[fila].height = 16

    # Resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"Resumen - {sitio_nombre}"
    ws2["A1"].font = Font(bold=True, size=13, color=AZUL)
    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A2"].font = Font(color="6B7280", size=9)

    for ci, t in enumerate(["Categoria", "Productos", "Con precio"], 1):
        c = ws2.cell(row=4, column=ci, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center")

    cnt = Counter(p["categoria"] for p in productos)
    cpr = Counter(p["categoria"] for p in productos if p["precio"] is not None)

    for ri, (cat, n) in enumerate(sorted(cnt.items()), 5):
        ws2.cell(row=ri, column=1, value=cat)
        ws2.cell(row=ri, column=2, value=n)
        ws2.cell(row=ri, column=3, value=cpr.get(cat, 0))

    tf = 5 + len(cnt)
    ws2.cell(row=tf, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=tf, column=2, value=len(productos)).font = Font(bold=True)
    ws2.cell(row=tf, column=3,
             value=sum(1 for p in productos if p["precio"] is not None)
             ).font = Font(bold=True)

    for col, w in [("A", 28), ("B", 14), ("C", 14)]:
        ws2.column_dimensions[col].width = w

    wb.save(nombre_f)
    print(f"  Excel: {nombre_f}")
    return nombre_f


# ==============================================================
# MAIN
# ==============================================================

async def main():
    sitio_arg = None
    if "--sitio" in sys.argv:
        idx = sys.argv.index("--sitio")
        if idx + 1 < len(sys.argv):
            sitio_arg = sys.argv[idx + 1].lower()

    if sitio_arg and sitio_arg not in SITIOS:
        print(f"Sitio no reconocido: '{sitio_arg}'")
        print(f"Disponibles: {', '.join(SITIOS.keys())}")
        sys.exit(1)

    sitios = [sitio_arg] if sitio_arg else list(SITIOS.keys())

    print(f"\nUbiqoCR Scraper v4")
    print(f"Sitios : {', '.join(sitios)}")
    print(f"Inicio : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    archivos = []
    for clave in sitios:
        cfg = SITIOS[clave]
        try:
            prods = await scrapear_sitio(cfg)
            if prods:
                ruta = exportar_excel(prods, cfg["nombre"])
                if ruta:
                    archivos.append(ruta)
        except KeyboardInterrupt:
            print("\nInterrumpido")
            break
        except Exception as e:
            print(f"\nError en {cfg['nombre']}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*60}")
    print(f"Fin: {datetime.now().strftime('%H:%M:%S')}")
    if archivos:
        print("Archivos:")
        for f in archivos:
            print(f"  - {f}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    asyncio.run(main())
