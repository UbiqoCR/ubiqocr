# UbiqoCR Scraper - EPA completo
# Todas las subcategorias de cr.epaenlinea.com
# Uso: python scraper.py --sitio epa

import asyncio
import re
import sys
import random
from datetime import datetime
from collections import Counter

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Instala: python -m pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instala: python -m pip install openpyxl")
    sys.exit(1)

CATEGORIAS_EPA = [
    # AUTOMOTRIZ
    ("https://cr.epaenlinea.com/accesorios-automotriz.html",     "Automotriz"),
    ("https://cr.epaenlinea.com/amarre.html",                     "Automotriz"),
    ("https://cr.epaenlinea.com/arrancadores-baterias.html",      "Automotriz"),
    ("https://cr.epaenlinea.com/baterias-automotriz.html",        "Automotriz"),
    ("https://cr.epaenlinea.com/bicicletas.html",                 "Automotriz"),
    ("https://cr.epaenlinea.com/cobertores-automotriz.html",      "Automotriz"),
    ("https://cr.epaenlinea.com/limpieza-automotriz.html",        "Automotriz"),
    ("https://cr.epaenlinea.com/lubricantes.html",                "Automotriz"),
    ("https://cr.epaenlinea.com/mecanica-automotriz.html",        "Automotriz"),
    ("https://cr.epaenlinea.com/seguridad-automotriz.html",       "Automotriz"),
    # BANOS
    ("https://cr.epaenlinea.com/accesorios-bano.html",            "Banos"),
    ("https://cr.epaenlinea.com/baneras.html",                    "Banos"),
    ("https://cr.epaenlinea.com/calentadores.html",               "Banos"),
    ("https://cr.epaenlinea.com/espejos.html",                    "Banos"),
    ("https://cr.epaenlinea.com/gabinetes-bano.html",             "Banos"),
    ("https://cr.epaenlinea.com/griferia-bano.html",              "Banos"),
    ("https://cr.epaenlinea.com/inodoros.html",                   "Banos"),
    ("https://cr.epaenlinea.com/lavamanos.html",                  "Banos"),
    ("https://cr.epaenlinea.com/muebles-bano.html",               "Banos"),
    ("https://cr.epaenlinea.com/regaderas.html",                  "Banos"),
    ("https://cr.epaenlinea.com/seguridad-bano.html",             "Banos"),
    # COCINAS
    ("https://cr.epaenlinea.com/accesorios-cocina.html",          "Cocinas"),
    ("https://cr.epaenlinea.com/filtros-agua.html",               "Cocinas"),
    ("https://cr.epaenlinea.com/fregaderos.html",                 "Cocinas"),
    ("https://cr.epaenlinea.com/griferia-cocina.html",            "Cocinas"),
    ("https://cr.epaenlinea.com/muebles-cocina.html",             "Cocinas"),
    ("https://cr.epaenlinea.com/sobres-de-cocina.html",           "Cocinas"),
    # DECORACION
    ("https://cr.epaenlinea.com/alfombras.html",                  "Decoracion"),
    ("https://cr.epaenlinea.com/almohadas.html",                  "Decoracion"),
    ("https://cr.epaenlinea.com/aromas.html",                     "Decoracion"),
    ("https://cr.epaenlinea.com/cojines.html",                    "Decoracion"),
    ("https://cr.epaenlinea.com/cortinas.html",                   "Decoracion"),
    ("https://cr.epaenlinea.com/cortineros.html",                 "Decoracion"),
    ("https://cr.epaenlinea.com/decoracion-paredes.html",         "Decoracion"),
    ("https://cr.epaenlinea.com/decoracion-ventanas.html",        "Decoracion"),
    ("https://cr.epaenlinea.com/frazadas.html",                   "Decoracion"),
    ("https://cr.epaenlinea.com/persianas.html",                  "Decoracion"),
    ("https://cr.epaenlinea.com/puertas-plegables.html",          "Decoracion"),
    ("https://cr.epaenlinea.com/ropa-cama.html",                  "Decoracion"),
    ("https://cr.epaenlinea.com/tapetes.html",                    "Decoracion"),
    # CONSTRUCCION
    ("https://cr.epaenlinea.com/aditivos-morteros.html",          "Construccion"),
    ("https://cr.epaenlinea.com/agregados.html",                  "Construccion"),
    ("https://cr.epaenlinea.com/aislantes.html",                  "Construccion"),
    ("https://cr.epaenlinea.com/alambres.html",                   "Construccion"),
    ("https://cr.epaenlinea.com/alcantarillas.html",              "Construccion"),
    ("https://cr.epaenlinea.com/angulares.html",                  "Construccion"),
    ("https://cr.epaenlinea.com/bloques-construccion.html",       "Construccion"),
    ("https://cr.epaenlinea.com/cal.html",                        "Construccion"),
    ("https://cr.epaenlinea.com/carretillas.html",                "Construccion"),
    ("https://cr.epaenlinea.com/cedazo.html",                     "Construccion"),
    ("https://cr.epaenlinea.com/cemento.html",                    "Construccion"),
    ("https://cr.epaenlinea.com/cielo-raso.html",                 "Construccion"),
    ("https://cr.epaenlinea.com/estereofon.html",                 "Construccion"),
    ("https://cr.epaenlinea.com/fibrocemento.html",               "Construccion"),
    ("https://cr.epaenlinea.com/gypsum.html",                     "Construccion"),
    ("https://cr.epaenlinea.com/hierro-forjado.html",             "Construccion"),
    ("https://cr.epaenlinea.com/impermeabilizantes-cementicios.html", "Construccion"),
    ("https://cr.epaenlinea.com/laminas-acrilicas.html",          "Construccion"),
    ("https://cr.epaenlinea.com/laminas-metalicas.html",          "Construccion"),
    ("https://cr.epaenlinea.com/laminas-techo.html",              "Construccion"),
    ("https://cr.epaenlinea.com/malla-electrosolada.html",        "Construccion"),
    ("https://cr.epaenlinea.com/malla-perimetral.html",           "Construccion"),
    ("https://cr.epaenlinea.com/morteros.html",                   "Construccion"),
    ("https://cr.epaenlinea.com/ocre.html",                       "Construccion"),
    ("https://cr.epaenlinea.com/pilas.html",                      "Construccion"),
    ("https://cr.epaenlinea.com/plasticos.html",                  "Construccion"),
    ("https://cr.epaenlinea.com/platinas.html",                   "Construccion"),
    ("https://cr.epaenlinea.com/rejillas-ventilacion.html",       "Construccion"),
    ("https://cr.epaenlinea.com/tejas.html",                      "Construccion"),
    ("https://cr.epaenlinea.com/tanques.html",                    "Construccion"),
    ("https://cr.epaenlinea.com/tubos-construccion.html",         "Construccion"),
    ("https://cr.epaenlinea.com/yeso.html",                       "Construccion"),
    ("https://cr.epaenlinea.com/varillas.html",                   "Construccion"),
    # ELECTRODOMESTICOS
    ("https://cr.epaenlinea.com/aires-acondicionados.html",       "Electrodomesticos"),
    ("https://cr.epaenlinea.com/aspiradoras.html",                "Electrodomesticos"),
    ("https://cr.epaenlinea.com/calentadores-ambiente.html",      "Electrodomesticos"),
    ("https://cr.epaenlinea.com/deshumidificadores.html",         "Electrodomesticos"),
    ("https://cr.epaenlinea.com/electrodomesticos-cocina.html",   "Electrodomesticos"),
    ("https://cr.epaenlinea.com/extractores-de-aire-bano.html",   "Electrodomesticos"),
    ("https://cr.epaenlinea.com/vaporizadores.html",              "Electrodomesticos"),
    ("https://cr.epaenlinea.com/ventiladores.html",               "Electrodomesticos"),
    # ELECTRICIDAD
    ("https://cr.epaenlinea.com/audio-y-video.html",              "Electricidad"),
    ("https://cr.epaenlinea.com/breakers.html",                   "Electricidad"),
    ("https://cr.epaenlinea.com/cables.html",                     "Electricidad"),
    ("https://cr.epaenlinea.com/cajas-electricidad.html",         "Electricidad"),
    ("https://cr.epaenlinea.com/casa-inteligente.html",           "Electricidad"),
    ("https://cr.epaenlinea.com/comunicacion.html",               "Electricidad"),
    ("https://cr.epaenlinea.com/centro-de-carga.html",            "Electricidad"),
    ("https://cr.epaenlinea.com/computadoras.html",               "Electricidad"),
    ("https://cr.epaenlinea.com/conexion-tierra.html",            "Electricidad"),
    ("https://cr.epaenlinea.com/empalmes.html",                   "Electricidad"),
    ("https://cr.epaenlinea.com/extensiones.html",                "Electricidad"),
    ("https://cr.epaenlinea.com/organizacion-cables.html",        "Electricidad"),
    ("https://cr.epaenlinea.com/multitomas.html",                 "Electricidad"),
    ("https://cr.epaenlinea.com/plaqueria.html",                  "Electricidad"),
    ("https://cr.epaenlinea.com/radios.html",                     "Electricidad"),
    ("https://cr.epaenlinea.com/router.html",                     "Electricidad"),
    ("https://cr.epaenlinea.com/sonido.html",                     "Electricidad"),
    ("https://cr.epaenlinea.com/sockets.html",                    "Electricidad"),
    ("https://cr.epaenlinea.com/televisores.html",                "Electricidad"),
    ("https://cr.epaenlinea.com/temporizadores.html",             "Electricidad"),
    ("https://cr.epaenlinea.com/terminales-electricos.html",      "Electricidad"),
    ("https://cr.epaenlinea.com/tuberia-metalica-electrica.html", "Electricidad"),
    ("https://cr.epaenlinea.com/tuberia-pvc-electrica.html",      "Electricidad"),
    # EXTERIORES
    ("https://cr.epaenlinea.com/accesorios-jardin.html",          "Exteriores"),
    ("https://cr.epaenlinea.com/botellas.html",                   "Exteriores"),
    ("https://cr.epaenlinea.com/calefactores-exterior.html",      "Exteriores"),
    ("https://cr.epaenlinea.com/camping.html",                    "Exteriores"),
    ("https://cr.epaenlinea.com/colchones-inflables.html",        "Exteriores"),
    ("https://cr.epaenlinea.com/combustibles.html",               "Exteriores"),
    ("https://cr.epaenlinea.com/control-de-plagas.html",          "Exteriores"),
    ("https://cr.epaenlinea.com/diversion.html",                  "Exteriores"),
    ("https://cr.epaenlinea.com/fertilizantes.html",              "Exteriores"),
    ("https://cr.epaenlinea.com/fuentes.html",                    "Exteriores"),
    ("https://cr.epaenlinea.com/hieleras.html",                   "Exteriores"),
    ("https://cr.epaenlinea.com/insecticidas.html",               "Exteriores"),
    ("https://cr.epaenlinea.com/macetas.html",                    "Exteriores"),
    ("https://cr.epaenlinea.com/mascotas.html",                   "Exteriores"),
    ("https://cr.epaenlinea.com/mobiliario-exterior.html",        "Exteriores"),
    ("https://cr.epaenlinea.com/paraguas.html",                   "Exteriores"),
    ("https://cr.epaenlinea.com/parrillas.html",                  "Exteriores"),
    ("https://cr.epaenlinea.com/piscinas.html",                   "Exteriores"),
    ("https://cr.epaenlinea.com/plantas-artificiales.html",       "Exteriores"),
    ("https://cr.epaenlinea.com/plantas-naturales.html",          "Exteriores"),
    ("https://cr.epaenlinea.com/riego.html",                      "Exteriores"),
    ("https://cr.epaenlinea.com/semillas.html",                   "Exteriores"),
    ("https://cr.epaenlinea.com/sombrillas.html",                 "Exteriores"),
    ("https://cr.epaenlinea.com/spa.html",                        "Exteriores"),
    ("https://cr.epaenlinea.com/sustratos.html",                  "Exteriores"),
    ("https://cr.epaenlinea.com/tierra.html",                     "Exteriores"),
    ("https://cr.epaenlinea.com/toldos.html",                     "Exteriores"),
    ("https://cr.epaenlinea.com/termos.html",                     "Exteriores"),
    # FERRETERIA Y CERRAJERIA
    ("https://cr.epaenlinea.com/abrazaderas.html",                "Ferreteria"),
    ("https://cr.epaenlinea.com/accesorios-cerrajeria.html",      "Ferreteria"),
    ("https://cr.epaenlinea.com/accesorios-muebles.html",         "Ferreteria"),
    ("https://cr.epaenlinea.com/agarraderas.html",                "Ferreteria"),
    ("https://cr.epaenlinea.com/anclajes.html",                   "Ferreteria"),
    ("https://cr.epaenlinea.com/arandelas.html",                  "Ferreteria"),
    ("https://cr.epaenlinea.com/barras-roscadas.html",            "Ferreteria"),
    ("https://cr.epaenlinea.com/cables-multiuso.html",            "Ferreteria"),
    ("https://cr.epaenlinea.com/cadenas.html",                    "Ferreteria"),
    ("https://cr.epaenlinea.com/candados.html",                   "Ferreteria"),
    ("https://cr.epaenlinea.com/cerraduras.html",                 "Ferreteria"),
    ("https://cr.epaenlinea.com/clavos.html",                     "Ferreteria"),
    ("https://cr.epaenlinea.com/cuerdas.html",                    "Ferreteria"),
    ("https://cr.epaenlinea.com/fijacion.html",                   "Ferreteria"),
    ("https://cr.epaenlinea.com/ganchos.html",                    "Ferreteria"),
    ("https://cr.epaenlinea.com/grapas.html",                     "Ferreteria"),
    ("https://cr.epaenlinea.com/portones.html",                   "Ferreteria"),
    ("https://cr.epaenlinea.com/proteccion-superficies.html",     "Ferreteria"),
    ("https://cr.epaenlinea.com/rodines.html",                    "Ferreteria"),
    ("https://cr.epaenlinea.com/ruedas.html",                     "Ferreteria"),
    ("https://cr.epaenlinea.com/tornillos.html",                  "Ferreteria"),
    ("https://cr.epaenlinea.com/tuercas.html",                    "Ferreteria"),
    # HERRAMIENTAS
    ("https://cr.epaenlinea.com/abrasivos.html",                  "Herramientas"),
    ("https://cr.epaenlinea.com/accesorios-herramientas.html",    "Herramientas"),
    ("https://cr.epaenlinea.com/alicates.html",                   "Herramientas"),
    ("https://cr.epaenlinea.com/bombas-de-agua.html",             "Herramientas"),
    ("https://cr.epaenlinea.com/brocas.html",                     "Herramientas"),
    ("https://cr.epaenlinea.com/cajas-herramientas.html",         "Herramientas"),
    ("https://cr.epaenlinea.com/cepillos-acero.html",             "Herramientas"),
    ("https://cr.epaenlinea.com/combos-herramientas.html",        "Herramientas"),
    ("https://cr.epaenlinea.com/compresores.html",                "Herramientas"),
    ("https://cr.epaenlinea.com/cubos.html",                      "Herramientas"),
    ("https://cr.epaenlinea.com/destornilladores.html",           "Herramientas"),
    ("https://cr.epaenlinea.com/discos-para-esmeril.html",        "Herramientas"),
    ("https://cr.epaenlinea.com/discos-para-sierras.html",        "Herramientas"),
    ("https://cr.epaenlinea.com/equipo-pintar.html",              "Herramientas"),
    ("https://cr.epaenlinea.com/esmeriles.html",                  "Herramientas"),
    ("https://cr.epaenlinea.com/generadores-electricos.html",     "Herramientas"),
    ("https://cr.epaenlinea.com/herramienta-carpinteria.html",    "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-agricola.html",      "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-albanileria.html",   "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-de-corte.html",      "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-de-fijacion.html",   "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-electricas-madera.html", "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-electricista.html",  "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-inalambricas.html",  "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-neumaticas.html",    "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-medicion.html",      "Herramientas"),
    ("https://cr.epaenlinea.com/herramientas-plomeria.html",      "Herramientas"),
    ("https://cr.epaenlinea.com/hidrolavadoras.html",             "Herramientas"),
    ("https://cr.epaenlinea.com/hidroneumaticos.html",            "Herramientas"),
    ("https://cr.epaenlinea.com/lijas.html",                      "Herramientas"),
    ("https://cr.epaenlinea.com/limas.html",                      "Herramientas"),
    ("https://cr.epaenlinea.com/llaves-de-apriete.html",          "Herramientas"),
    ("https://cr.epaenlinea.com/maquinaria-jardin.html",          "Herramientas"),
    ("https://cr.epaenlinea.com/martillos.html",                  "Herramientas"),
    ("https://cr.epaenlinea.com/mototools.html",                  "Herramientas"),
    ("https://cr.epaenlinea.com/prensas-manuales.html",           "Herramientas"),
    ("https://cr.epaenlinea.com/pulidoras.html",                  "Herramientas"),
    ("https://cr.epaenlinea.com/soldadura.html",                  "Herramientas"),
    ("https://cr.epaenlinea.com/taladros.html",                   "Herramientas"),
    ("https://cr.epaenlinea.com/tronzadoras.html",                "Herramientas"),
    # LAMPARAS
    ("https://cr.epaenlinea.com/apliques.html",                   "Lamparas"),
    ("https://cr.epaenlinea.com/bombillos.html",                  "Lamparas"),
    ("https://cr.epaenlinea.com/faroles.html",                    "Lamparas"),
    ("https://cr.epaenlinea.com/iluminacion-exterior.html",       "Lamparas"),
    ("https://cr.epaenlinea.com/iluminacion-solar.html",          "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-auxiliares.html",        "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-colgantes.html",         "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-comerciales.html",       "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-de-pie.html",            "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-empotrables.html",       "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-escritorio.html",        "Lamparas"),
    ("https://cr.epaenlinea.com/lamparas-mesa.html",              "Lamparas"),
    ("https://cr.epaenlinea.com/plafones.html",                   "Lamparas"),
    ("https://cr.epaenlinea.com/ventiladores-techo.html",         "Lamparas"),
    # LIMPIEZA
    ("https://cr.epaenlinea.com/articulos-de-limpieza.html",      "Limpieza"),
    ("https://cr.epaenlinea.com/control-de-olores.html",          "Limpieza"),
    ("https://cr.epaenlinea.com/desatoradores.html",              "Limpieza"),
    ("https://cr.epaenlinea.com/desinfectantes.html",             "Limpieza"),
    ("https://cr.epaenlinea.com/lavanderia.html",                 "Limpieza"),
    ("https://cr.epaenlinea.com/liquidos-limpieza.html",          "Limpieza"),
    ("https://cr.epaenlinea.com/planchadores.html",               "Limpieza"),
    # MADERAS Y PUERTAS
    ("https://cr.epaenlinea.com/aglomerados.html",                "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/laminado-melaminico.html",        "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/machihembrados.html",             "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/marcos-para-puertas.html",        "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/molduras.html",                   "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/plywood.html",                    "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/puertas-interior.html",           "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/puertas-closet.html",             "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/reglas-madera.html",              "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/tableros-madera.html",            "Maderas y Puertas"),
    ("https://cr.epaenlinea.com/ventanas.html",                   "Maderas y Puertas"),
    # MUEBLES Y ORGANIZACION
    ("https://cr.epaenlinea.com/armarios.html",                   "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/bases-tv.html",                   "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/basureros.html",                  "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/cajas.html",                      "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/carros-de-mercado.html",          "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/cestas-organizadoras.html",       "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/closet.html",                     "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/colchones.html",                  "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/estantes.html",                   "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/ganchos-para-ropa.html",          "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/gaveteros.html",                  "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/muebles-interior.html",           "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/mesas-interior.html",             "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/organizadores.html",              "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/otomanes.html",                   "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/percheros.html",                  "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/repisas.html",                    "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/sillas-interior.html",            "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/sofas.html",                      "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/soportes-repisas.html",           "Muebles y Organizacion"),
    ("https://cr.epaenlinea.com/zapateras.html",                  "Muebles y Organizacion"),
    # PINTURAS
    ("https://cr.epaenlinea.com/acabado-para-maderas.html",       "Pinturas"),
    ("https://cr.epaenlinea.com/accesorios-pintar.html",          "Pinturas"),
    ("https://cr.epaenlinea.com/anticorrosivos.html",             "Pinturas"),
    ("https://cr.epaenlinea.com/cintas.html",                     "Pinturas"),
    ("https://cr.epaenlinea.com/escaleras.html",                  "Pinturas"),
    ("https://cr.epaenlinea.com/impermeabilizacion-superficies.html", "Pinturas"),
    ("https://cr.epaenlinea.com/pegamentos.html",                 "Pinturas"),
    ("https://cr.epaenlinea.com/pinturas.html",                   "Pinturas"),
    ("https://cr.epaenlinea.com/pintura-spray.html",              "Pinturas"),
    ("https://cr.epaenlinea.com/reparacion-superficies.html",     "Pinturas"),
    ("https://cr.epaenlinea.com/revestimientos.html",             "Pinturas"),
    ("https://cr.epaenlinea.com/selladores-adhesivos.html",       "Pinturas"),
    ("https://cr.epaenlinea.com/silicon.html",                    "Pinturas"),
    ("https://cr.epaenlinea.com/solventes.html",                  "Pinturas"),
    # PISOS
    ("https://cr.epaenlinea.com/bloques-de-vidrio.html",          "Pisos"),
    ("https://cr.epaenlinea.com/ceramica-para-piso.html",         "Pisos"),
    ("https://cr.epaenlinea.com/ceramica-pared-azulejos.html",    "Pisos"),
    ("https://cr.epaenlinea.com/fachaletas.html",                 "Pisos"),
    ("https://cr.epaenlinea.com/fraguas.html",                    "Pisos"),
    ("https://cr.epaenlinea.com/mallas-para-pared.html",          "Pisos"),
    ("https://cr.epaenlinea.com/paneles-decorativos.html",        "Pisos"),
    ("https://cr.epaenlinea.com/pisos-laminados.html",            "Pisos"),
    ("https://cr.epaenlinea.com/pisos-tipo-deck.html",            "Pisos"),
    ("https://cr.epaenlinea.com/porcelanato.html",                "Pisos"),
    ("https://cr.epaenlinea.com/pisos-vinilicos.html",            "Pisos"),
    ("https://cr.epaenlinea.com/rodapies.html",                   "Pisos"),
    # PLOMERIA
    ("https://cr.epaenlinea.com/accesorios-inodoro.html",         "Plomeria"),
    ("https://cr.epaenlinea.com/canoas.html",                     "Plomeria"),
    ("https://cr.epaenlinea.com/desagues.html",                   "Plomeria"),
    ("https://cr.epaenlinea.com/kits-instalacion-plomeria.html",  "Plomeria"),
    ("https://cr.epaenlinea.com/llave-chorro.html",               "Plomeria"),
    ("https://cr.epaenlinea.com/mangueras.html",                  "Plomeria"),
    ("https://cr.epaenlinea.com/pegamentos-tuberia.html",         "Plomeria"),
    ("https://cr.epaenlinea.com/repuestos-griferia.html",         "Plomeria"),
    ("https://cr.epaenlinea.com/trampas-de-grasa.html",           "Plomeria"),
    ("https://cr.epaenlinea.com/tuberia-galvanizada.html",        "Plomeria"),
    ("https://cr.epaenlinea.com/tuberia-gas.html",                "Plomeria"),
    ("https://cr.epaenlinea.com/tuberia-pvc.html",                "Plomeria"),
    ("https://cr.epaenlinea.com/tubos-abasto.html",               "Plomeria"),
    ("https://cr.epaenlinea.com/valvulas-plomeria.html",          "Plomeria"),
    # SEGURIDAD
    ("https://cr.epaenlinea.com/alarmas.html",                    "Seguridad"),
    ("https://cr.epaenlinea.com/baterias.html",                   "Seguridad"),
    ("https://cr.epaenlinea.com/cajas-de-seguridad.html",         "Seguridad"),
    ("https://cr.epaenlinea.com/camaras-de-seguridad.html",       "Seguridad"),
    ("https://cr.epaenlinea.com/equipo-proteccion.html",          "Seguridad"),
    ("https://cr.epaenlinea.com/extintores.html",                 "Seguridad"),
    ("https://cr.epaenlinea.com/lamparas-emergencia.html",        "Seguridad"),
    ("https://cr.epaenlinea.com/linternas.html",                  "Seguridad"),
    ("https://cr.epaenlinea.com/porteros.html",                   "Seguridad"),
    ("https://cr.epaenlinea.com/seguridad-infantil.html",         "Seguridad"),
    ("https://cr.epaenlinea.com/senalizacion.html",               "Seguridad"),
    ("https://cr.epaenlinea.com/timbres.html",                    "Seguridad"),
    ("https://cr.epaenlinea.com/veladoras.html",                  "Seguridad"),
]

SITIO = {
    "nombre": "EPA",
    "categorias": CATEGORIAS_EPA,
    "selector_prod":   ".product-item-info",
    "selector_nombre": ".product-item-name a",
    "selector_precio": "span.price",
    "selector_sig":    "a.action.next",
    "paginacion":      "?p=",
}


def limpiar_precio(texto):
    if not texto:
        return None
    limpio = re.sub(r'[^\d.,]', '', texto.strip())
    if not limpio:
        return None
    tiene_punto = '.' in limpio
    tiene_coma  = ',' in limpio
    if tiene_punto and tiene_coma:
        if limpio.rfind('.') > limpio.rfind(','):
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace('.', '').replace(',', '.')
    elif tiene_punto:
        partes = limpio.split('.')
        if len(partes) >= 2 and len(partes[-1]) == 3:
            limpio = limpio.replace('.', '')
        elif len(partes) > 2:
            limpio = limpio.replace('.', '')
    elif tiene_coma:
        partes = limpio.split(',')
        if len(partes) == 2 and len(partes[1]) == 3:
            limpio = limpio.replace(',', '')
        else:
            limpio = limpio.replace(',', '.')
    try:
        r = float(limpio)
        return r if r > 0 else None
    except ValueError:
        return None


def limpiar_texto(s):
    if not isinstance(s, str):
        return s
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s).strip()


def url_pag(url_base, pagina):
    if pagina == 1:
        return url_base
    sep = "&" if "?" in url_base else "?"
    return url_base + sep + "p=" + str(pagina)


async def scrapear():
    cfg       = SITIO
    productos = []
    vistos    = set()

    print(f"\n{'='*60}")
    print(f"  EPA - {len(cfg['categorias'])} subcategorias")
    print(f"{'='*60}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox",
                  "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
            locale="es-CR",
        )
        page = await context.new_page()
        await page.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,woff,woff2,ttf,eot,ico}",
            lambda route: route.abort()
        )

        for url_cat, cat_nombre in cfg["categorias"]:
            pagina = 1
            errores = 0
            print(f"\n  [{cat_nombre}] {url_cat.split('/')[-1]}")

            while pagina <= 100:
                url = url_pag(url_cat, pagina)
                print(f"    p{pagina:>3} -> ", end="", flush=True)

                try:
                    resp = await page.goto(url, wait_until="domcontentloaded", timeout=25000)
                    if resp and resp.status in (404, 403, 410):
                        print(f"HTTP {resp.status}")
                        break

                    try:
                        await page.wait_for_selector(cfg["selector_prod"], timeout=10000)
                    except PWTimeout:
                        print("sin productos")
                        errores += 1
                        if errores >= 3:
                            break
                        continue

                except PWTimeout:
                    print("timeout")
                    errores += 1
                    if errores >= 3:
                        break
                    await asyncio.sleep(2)
                    continue
                except Exception as e:
                    print(f"err: {str(e)[:40]}")
                    errores += 1
                    if errores >= 3:
                        break
                    continue

                errores = 0
                items = await page.query_selector_all(cfg["selector_prod"])
                if not items:
                    print("vacio")
                    break

                nuevos = 0
                for item in items:
                    try:
                        el_n = await item.query_selector(cfg["selector_nombre"])
                        if not el_n:
                            continue
                        nombre = limpiar_texto((await el_n.inner_text()).strip())
                        if not nombre:
                            continue

                        precio_txt = ""
                        el_p = await item.query_selector(cfg["selector_precio"])
                        if el_p:
                            precio_txt = (await el_p.inner_text()).strip()
                        precio = limpiar_precio(precio_txt)

                        href = limpiar_texto(await el_n.get_attribute("href") or "")

                        imagen = ""
                        try:
                            img_el = await item.query_selector("img")
                            if img_el:
                                imagen = limpiar_texto(
                                    await img_el.get_attribute("src") or
                                    await img_el.get_attribute("data-src") or ""
                                )
                        except Exception:
                            pass

                        clave = nombre.lower()
                        if clave in vistos:
                            continue
                        vistos.add(clave)

                        productos.append({
                            "nombre":    nombre,
                            "precio":    precio,
                            "moneda":    "CRC",
                            "categoria": cat_nombre,
                            "url":       href,
                            "imagen":    imagen,
                        })
                        nuevos += 1
                    except Exception:
                        continue

                print(f"{nuevos:>3} nuevos | total: {len(productos):>6}")

                sig = await page.query_selector(cfg["selector_sig"])
                if not sig:
                    break
                pagina += 1
                await asyncio.sleep(random.uniform(0.8, 1.8))

        await browser.close()

    sin_p = sum(1 for p in productos if p["precio"] is None)
    print(f"\n  TOTAL EPA: {len(productos)} productos ({sin_p} sin precio)")
    return productos


def exportar_excel(productos):
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_EPA_{fecha}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    AZUL  = "2563EB"
    CLARO = "EFF6FF"
    GRIS  = "F8FAFC"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hx = PatternFill("solid", fgColor=AZUL)
    ha = Alignment(horizontal="center", vertical="center")
    ax = PatternFill("solid", fgColor=CLARO)
    gx = PatternFill("solid", fgColor=GRIS)
    na = Alignment(vertical="center")
    ra = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="E2E8F0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols   = ["Nombre del producto", "Precio", "Moneda", "Categoria", "URL", "Imagen URL"]
    widths = [55, 14, 9, 25, 60, 60]

    for ci, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = hf; cell.fill = hx
        cell.alignment = ha; cell.border = brd
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for i, prod in enumerate(productos, 1):
        fila = i + 1
        fill = ax if i % 2 == 0 else gx
        vals = [prod.get("nombre",""), prod.get("precio"),
                prod.get("moneda","CRC"), prod.get("categoria",""), prod.get("url",""),
                prod.get("imagen","")]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, str):
                v = limpiar_texto(v)
            cell = ws.cell(row=fila, column=ci, value=v)
            cell.border = brd; cell.fill = fill
            if ci == 2 and isinstance(v, float):
                cell.number_format = "#,##0.00"; cell.alignment = ra
            else:
                cell.alignment = na
        ws.row_dimensions[fila].height = 16

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = f"EPA - {len(productos)} productos"
    ws2["A1"].font = Font(bold=True, size=13, color=AZUL)
    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A2"].font = Font(color="6B7280", size=9)

    for ci, t in enumerate(["Categoria", "Productos", "Con precio"], 1):
        c = ws2.cell(row=4, column=ci, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)

    cnt = Counter(p["categoria"] for p in productos)
    cpr = Counter(p["categoria"] for p in productos if p["precio"] is not None)
    for ri, (cat, n) in enumerate(sorted(cnt.items()), 5):
        ws2.cell(row=ri, column=1, value=cat)
        ws2.cell(row=ri, column=2, value=n)
        ws2.cell(row=ri, column=3, value=cpr.get(cat, 0))

    tf = 5 + len(cnt)
    ws2.cell(row=tf, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=tf, column=2, value=len(productos)).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(nombre_f)
    print(f"\n  Excel guardado: {nombre_f}")
    return nombre_f


def exportar_json(productos):
    import json
    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_f = f"productos_EPA_{fecha}.json"
    data = [{"negocio": "EPA", **p} for p in productos]
    with open(nombre_f, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON guardado:  {nombre_f}")
    return nombre_f


async def main():
    print(f"\nUbiqoCR Scraper - EPA completo")
    print(f"Inicio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"Total subcategorias: {len(CATEGORIAS_EPA)}")

    prods = await scrapear()
    if prods:
        exportar_excel(prods)
        exportar_json(prods)

    print(f"\nFin: {datetime.now().strftime('%H:%M:%S')}\n")


if __name__ == "__main__":
    asyncio.run(main())
